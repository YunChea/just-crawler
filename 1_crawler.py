import time
from requests.adapters import HTTPAdapter
from bs4 import BeautifulSoup
import requests
import urllib3
urllib3.disable_warnings()
from lxml import etree
import openpyxl
headers = {'user-agent': "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36",}
url1 = 'https://nynct.henan.gov.cn/qcjgxq/index.html'
url2 = 'https://nynct.henan.gov.cn/qcjgxq/index_'
k = input()
def crawler(url):
    list_page = requests.get(url = url1,headers = headers,verify=False)
    list_page.keep_alive = False
    list_page.encoding = 'utf-8'
    list_pagetext = list_page.text
    soup = BeautifulSoup(list_pagetext, 'lxml')
    for urls in soup.find_all('li'):
        wb_name = urls.text
        urls = urls.select('a')[0]
        detail_url = urls.get('href')
        print(detail_url,wb_name)
        page = requests.get(url=detail_url, headers=headers,verify=False)
        page.keep_alive = False
        page.encoding = 'utf-8'
        pagetext = page.text
        soup = BeautifulSoup(pagetext, 'lxml')
        product = soup.find_all('td')
        i = 1
        #wb = openpyxl.load_workbook(wb_name+'.xlsx')
        wb = openpyxl.Workbook()
        #sheet = wb['Sheet1']
        wb.create_sheet('Sheet1')
        ws = wb.active
        j = 1
        for td in product:
            ws.cell(j, i, td.text)
            i = i + 1
            if i % 10 == 0:
                i = 1
                j = j + 1
        wb.save(wb_name+'.xlsx')
        time.sleep(5)
for i in range(int(k)+1):
    if i == 0:
        url = url1
        print(url)
        crawler(url)
    else:
        url = url2+str(i)+'.html'
        print(url)
        crawler(url)